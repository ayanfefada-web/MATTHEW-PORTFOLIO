# ============================================================
#  TASK 1: Web Scraping
#  Name: Oyeyiola
#  Tool: BeautifulSoup + Requests + Pandas
#  Website: books.toscrape.com
#  Output: books_dataset.xlsx + books_dataset.csv
# ============================================================
#
#  In this task, I used Python's BeautifulSoup library to
#  scrape book data from a public website (books.toscrape.com).
#  I extracted the title, price, rating, and availability of
#  each book, then saved the results into a clean dataset
#  for further analysis.
#
#  HOW TO RUN:
#  1. Open Command Prompt or Terminal
#  2. Install the required libraries (only once):
#       pip install requests beautifulsoup4 pandas openpyxl
#  3. Run this script:
#       python task1_web_scraper.py
#  4. Output files will appear in the same folder
# ============================================================

# Step 1: I imported the libraries I needed
import requests
from bs4 import BeautifulSoup
import pandas as pd
import time

# Step 2: I defined the website I wanted to scrape
BASE_URL = "https://books.toscrape.com/catalogue/page-{}.html"
TOTAL_PAGES = 10  # I chose 10 pages to get 200 books

# I mapped the word ratings to numbers for easier analysis
RATING_MAP = {
    "One": 1,
    "Two": 2,
    "Three": 3,
    "Four": 4,
    "Five": 5
}

# Step 3: I wrote a function to scrape one page at a time
def scrape_page(page_number):
    url = BASE_URL.format(page_number)
    response = requests.get(url, timeout=10)

    if response.status_code != 200:
        print(f"  [!] Could not load page {page_number} — Status: {response.status_code}")
        return []

    # I used BeautifulSoup to parse the HTML of the page
    soup = BeautifulSoup(response.text, "html.parser")
    books = []

    # I looped through each book card on the page
    for book in soup.find_all("article", class_="product_pod"):
        title = book.h3.a["title"]
        price_text = book.find("p", class_="price_color").text.strip()
        price = float(price_text.replace("£", "").replace("Â", ""))
        rating_word = book.p["class"][1]
        rating = RATING_MAP.get(rating_word, 0)
        availability = book.find("p", class_="instock availability").text.strip()

        books.append({
            "Title": title,
            "Price (£)": price,
            "Rating (out of 5)": rating,
            "Availability": availability,
            "Scraped From": f"books.toscrape.com — Page {page_number}"
        })

    return books


# Step 4: I added extra columns to make the dataset more useful
def add_derived_columns(df):
    df["Price Category"] = pd.cut(
        df["Price (£)"],
        bins=[0, 20, 35, 50, 100],
        labels=["Budget (<£20)", "Mid (£20–35)", "Standard (£35–50)", "Premium (£50+)"]
    )
    df["Star Label"] = df["Rating (out of 5)"].apply(
        lambda r: "★" * r + "☆" * (5 - r)
    )
    return df


# Step 5: I formatted and saved the data into a professional Excel file
def save_to_excel(df, filename="books_dataset.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    DARK = '1F3864'; BLUE = '2E75B6'; LBLUE = 'BDD7EE'
    LGREY = 'F2F2F2'; WHITE = 'FFFFFF'

    def border():
        s = Side(style='thin', color='BFBFBF')
        return Border(left=s, right=s, top=s, bottom=s)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Scraped Books Data'
    ws.sheet_view.showGridLines = False

    cols = list(df.columns)
    col_widths = [55, 14, 18, 16, 30, 18, 14]
    ws.row_dimensions[1].height = 28

    for i, (col, w) in enumerate(zip(cols, col_widths), 1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = Font(name='Arial', bold=True, color=WHITE, size=11)
        c.fill = PatternFill('solid', start_color=DARK)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(i)].width = w

    for r, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        bg = LGREY if r % 2 == 0 else WHITE
        ws.row_dimensions[r].height = 18
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=r, column=ci, value=str(val) if ci in [4, 6, 7] else val)
            c.font = Font(name='Arial', size=10)
            c.fill = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(horizontal='center' if ci != 1 else 'left', vertical='center')
            c.border = border()
            if ci == 2:
                c.number_format = '£#,##0.00'

    ws.freeze_panes = 'A2'
    wb.save(filename)
    print(f"  [✓] Excel file saved: {filename}")


# Step 6: I ran everything together in the main function
def main():
    print("=" * 55)
    print("  TASK 1: Web Scraping — books.toscrape.com")
    print("=" * 55)

    all_books = []

    for page in range(1, TOTAL_PAGES + 1):
        print(f"  Scraping page {page} of {TOTAL_PAGES}...")
        books = scrape_page(page)
        all_books.extend(books)
        time.sleep(0.5)  # I added a small delay to be respectful to the server

    print(f"\n  Total books scraped: {len(all_books)}")

    df = pd.DataFrame(all_books)
    df = add_derived_columns(df)

    # I saved the dataset in two formats
    df.to_csv("books_dataset.csv", index=False)
    print("  [✓] CSV file saved: books_dataset.csv")

    save_to_excel(df, "books_dataset.xlsx")

    # I printed a final summary of what I collected
    print("\n" + "=" * 55)
    print("  SUMMARY OF MY SCRAPED DATASET")
    print("=" * 55)
    print(f"  Total Books Scraped : {len(df)}")
    print(f"  Average Price       : £{df['Price (£)'].mean():.2f}")
    print(f"  Average Rating      : {df['Rating (out of 5)'].mean():.1f} / 5")
    print(f"  In Stock            : {len(df[df['Availability'] == 'In stock'])}")
    print(f"  Out of Stock        : {len(df[df['Availability'] != 'In stock'])}")
    print("=" * 55)
    print("\n  Done! Open books_dataset.xlsx to view my dataset.\n")


if __name__ == "__main__":
    main()
